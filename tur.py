from openpyxl import workbook, load_workbook
import openpyxl
from flask import Flask, render_template, request, redirect, url_for, session,g
import re
import distance
import google.generativeai as genai
import subprocess
from difflib import get_close_matches
genai.configure(api_key='AIzaSyCco1maX_IrGKMwf8r1U45DDtzFRWryY1M')
model = genai.GenerativeModel('gemini-pro')
a=['Sweden', 'Finland', 'Austria', 'Canada', 'Switzerland', 'United Kingdom', 'United States of America', 'Cyprus', 'Mauritius', 'Malaysia', 'New Zealand', 'Ireland', 'Singapore', 'Dubai', 'France', 'Netherlands', 'Hungary', 'Spain', 'Italy', 'Denmark', 'Lithuania', 'Japan', 'Malta', 'Vietnam', 'Poland', 'China', 'Russia', 'Belgium', 'South Korea', 'Georgia', 'India', 'Australia']
b=[]
app = Flask(__name__, template_folder="..\\ne2")
@app.route('/')
@app.route("/country", methods=['GET', 'POST'])
def country():
    if request.method == 'POST':
            if request.form.get('pinch') == 'next':
                   x=request.form.getlist('mycheck')
                   for i in x:
                        b.append(i)
                   print(x,"--->",type(x))
                   return redirect("/intake")
            else:
                  return render_template("ir.html",ab=a)
    elif request.method == 'GET':
        print("No Post Back Call")
    return render_template("ir.html",ab=a)
@app.route("/intake", methods=['GET', 'POST'])
def intake():
        if request.method == 'POST':
                global option
                if request.form.get('eel') == 'Next':
                      option = request.form['country']
                      option=str(option)
                      print(option,"--->",type(option))
                      return redirect("/process")
                else:
                    return render_template("index.html")  


        elif request.method == 'GET':
            print("No Post Back Call")
        return render_template("index.html")
@app.route('/process', methods=['GET','POST']) 
def process():
    if request.method=='POST':
        if(len(option)>0):
            global name
            global age
            global height
            global w
            json_data = request.get_json()
            name = json_data.get('name')
            age = json_data.get('age')
            height=json_data.get('height')
            weight=json_data.get('weight')
            w=str(weight)
            print(age)
            print(name)
            print(height)
            print(w)

    else:
        return render_template("nas.html") 
    return render_template("nas.html")
   
@app.route('/ossa', methods=['GET','POST']) 
def ossa():
    global ie
    global to
    ie=6.5
    to=90
    if request.method == 'POST':
         xx=request.form.get('exam')
         print(xx)
    elif request.method == 'GET':
        print("No Post Back Call")
    return render_template("oya.html")
@app.route('/ossa1', methods=['GET','POST']) 
def ossa1():
    if request.method == 'POST':
           global money
           if(request.form.get('su')=='prosceed'):
                money=request.form.get('tu')    
                money=int(money)
                return redirect('/final')
           elif(request.form.get('su')=='next'):
                xv=request.form.get('mycheck')
                print(xv)
                money=151000
                #wb = load_workbook("D:\\details\\ne2\\jamey.xlsx")
                #ws = wb.active
                jil=[]
                #jil.append('#1474')
                #jil.append('abc@gmail.com')
                #jil.append('8939200603')
                #jil.append(age)
                #jil.append(name)
                #jil.append(ie)
                #jil.append(to)
                #ws.append(jil)
                #wb.save("D:\\details\\ne2\\jamey.xlsx")
                return redirect('/final')
    elif request.method == 'GET':
        print("No Post Back Call")
    return render_template("mon.html")
"""PG Diploma /Certificate-age
MSc Biomedical Sciences Research-name
AL-height
Fulltime-w"""
"""PG Diploma /Certificate-age
MSc Biomedical Sciences Research-name
AL-height
Fulltime-w"""
@app.route('/final', methods=['GET','POST']) 
def final():
     ie=6.5
     to=90
     workbook = openpyxl.load_workbook("C:\\Users\\Sanchith Abinav\\Downloads\\University Data Ali Global Education.xlsm") 
     sheet = workbook.worksheets[0]
     zep=[]
     uni=[]
     for i in sheet:
          for j in b:
               if(i[21].value!='ielts_overall' and i[21].value!=None and i[22].value!='toefl_score' and i[22].value!=None):
                    ie1=int(i[21].value)
                    to1=int(i[22].value)
                    print(ie1,"--->",to1)
               if(i[32].value!='amount'):
                           am=int(i[32].value)
                           print(am)
               if(i[5].value==j and i[4].value==age and ie1<=ie and to1<=to and am<=money):
                    print(i[1].value)
                    zep.append(i[1].value)
     xam=get_close_matches(name,zep)
     print(xam)
     styleo=[]
     lala={}
     for i in sheet:
          for p,k in zip(xam,b):
                    if(i[1].value==p and i[5].value==k):
                         print(i[6].value)
                         uni.append(i[6].value)

     global s
     if request.method=='POST':
                    e = request.form.get('button_id')
                    e=str(e)
                    if(e=='courses'):
                          o=request.form.get('button_value')
                          o=str(o)
                          pr="list out all the courses that are available in"+o+" "+"according to top universities.com"
                          response = model.generate_content(pr)
                          s=response.text
                          #pol="python t.py"+" "+s
                          #subprocess.run(pol)
                          return render_template("final1.html",alcapone=s)
                    elif(e=='scholar'):
                          o=request.form.get('button_value')
                          o=str(o)
                          pr="list out all the scholarships that are available in"+o+" "+"according to top universities.com"
                          response = model.generate_content(pr)
                          s=response.text
                          #pol="python t.py"+" "+s
                          #subprocess.run(pol)
                          return render_template("final1.html",alcapone=s)

                    elif(e=='fee'):
                          o=request.form.get('button_value')
                          o=str(o)
                          pr="what's the fee structure for each semester in"+o+" "+"according to top universities.com give me result for each semester"
                          response = model.generate_content(pr)
                          s=response.text
                          #pol="python t.py"+" "+s
                          #subprocess.run(pol)
                          return render_template("final1.html",alcapone=s)

                    elif(e=='carrear'):
                          o=request.form.get('button_value')
                          o=str(o)
                          pr="which companies come to recruit the people that study"+name+"at"+o+"and suggest some carrear opportunities with minumum package"
                          response = model.generate_content(pr)
                          s=response.text
                          print(s)
                          #pol="python t.py"+" "+s
                          #subprocess.run(pol)
                          return render_template("final1.html",alcapone=s)

     else:
           return render_template("final.html",x0=uni,y0=xam,z0=b )
                   
     return render_template("final.html",x0=uni,y0=xam,z0=b,f=age,f0=name)

           
          
if __name__ == '__main__':
    app.run(debug=False)

